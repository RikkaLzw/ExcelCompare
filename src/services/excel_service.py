"""
Excel 文件解析服务

支持 .xlsx 和 .xls 格式的文件读取。
"""
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

from src.models.excel_model import (
    CellData, CellStyle, CellType, SheetData, WorkbookData
)


class ExcelService:
    """Excel 文件解析服务"""
    
    # 支持的文件扩展名
    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls'}
    # 最大文件大小 (100MB)
    MAX_FILE_SIZE = 100 * 1024 * 1024
    
    @classmethod
    def load_file(cls, file_path: str) -> WorkbookData:
        """
        加载 Excel 文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            WorkbookData 对象
            
        Raises:
            ValueError: 文件格式不支持或文件无效
            FileNotFoundError: 文件不存在
        """
        path = Path(file_path)
        
        # 检查文件是否存在
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        # 检查文件扩展名
        ext = path.suffix.lower()
        if ext not in cls.SUPPORTED_EXTENSIONS:
            raise ValueError(f"不支持的文件格式: {ext}，仅支持 .xlsx 和 .xls")
        
        # 检查文件大小
        file_size = path.stat().st_size
        if file_size > cls.MAX_FILE_SIZE:
            raise ValueError(f"文件大小超过限制 (最大 100MB)")
        
        # 获取文件信息
        modified_time = datetime.fromtimestamp(path.stat().st_mtime)
        
        # 根据扩展名选择解析方法
        if ext == '.xlsx':
            sheets = cls._load_xlsx(file_path)
        else:  # .xls
            sheets = cls._load_xls(file_path)
        
        return WorkbookData(
            file_path=file_path,
            file_name=path.name,
            file_size=file_size,
            modified_time=modified_time.strftime("%Y-%m-%d %H:%M:%S"),
            sheets=sheets,
            sheet_names=[s.name for s in sheets]
        )
    
    @classmethod
    def _load_xlsx(cls, file_path: str) -> list[SheetData]:
        """加载 .xlsx 文件"""
        try:
            # data_only=True 获取计算后的值，而不是公式
            wb = openpyxl.load_workbook(file_path, data_only=False, read_only=False)
            sheets = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                
                # 获取实际使用的范围
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                
                for row_idx in range(1, max_row + 1):
                    row_data = []
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell_data = cls._parse_cell(cell)
                        row_data.append(cell_data)
                    rows.append(row_data)
                
                sheets.append(SheetData(
                    name=sheet_name,
                    rows=rows,
                    row_count=max_row,
                    col_count=max_col
                ))
            
            wb.close()
            return sheets
            
        except Exception as e:
            raise ValueError(f"无法读取 Excel 文件: {str(e)}")
    
    @classmethod
    def _load_xls(cls, file_path: str) -> list[SheetData]:
        """加载 .xls 文件（使用 xlrd）"""
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sheets = []
            
            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                rows = []
                
                for row_idx in range(ws.nrows):
                    row_data = []
                    for col_idx in range(ws.ncols):
                        cell = ws.cell(row_idx, col_idx)
                        cell_data = cls._parse_xls_cell(cell, wb)
                        row_data.append(cell_data)
                    rows.append(row_data)
                
                sheets.append(SheetData(
                    name=ws.name,
                    rows=rows,
                    row_count=ws.nrows,
                    col_count=ws.ncols
                ))
            
            return sheets
            
        except ImportError:
            raise ValueError("需要安装 xlrd 库来读取 .xls 文件")
        except Exception as e:
            raise ValueError(f"无法读取 Excel 文件: {str(e)}")
    
    @classmethod
    def _parse_cell(cls, cell: Cell) -> CellData:
        """解析 openpyxl 单元格"""
        value = cell.value
        formula = None
        cell_type = CellType.EMPTY
        
        # 检查公式
        if cell.data_type == 'f' or (isinstance(value, str) and value.startswith('=')):
            formula = str(value) if isinstance(value, str) and value.startswith('=') else None
            cell_type = CellType.FORMULA
        elif value is None or value == "":
            cell_type = CellType.EMPTY
        elif isinstance(value, bool):
            cell_type = CellType.BOOLEAN
        elif isinstance(value, (int, float)):
            cell_type = CellType.NUMBER
        elif isinstance(value, datetime):
            cell_type = CellType.DATE
        else:
            cell_type = CellType.STRING
        
        # 解析样式
        style = None
        if cell.font or cell.fill or cell.border:
            style = CellStyle(
                font_name=cell.font.name if cell.font else None,
                font_size=cell.font.size if cell.font else None,
                font_bold=cell.font.bold if cell.font else False,
                font_italic=cell.font.italic if cell.font else False,
                font_color=str(cell.font.color.rgb) if cell.font and cell.font.color and cell.font.color.rgb else None,
                bg_color=str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb else None,
                number_format=cell.number_format
            )
        
        # 解析批注
        comment = None
        if cell.comment:
            comment = cell.comment.text
        
        return CellData(
            value=value,
            formula=formula,
            cell_type=cell_type,
            style=style,
            comment=comment
        )
    
    @classmethod
    def _parse_xls_cell(cls, cell, workbook) -> CellData:
        """解析 xlrd 单元格"""
        import xlrd
        
        value = cell.value
        cell_type = CellType.EMPTY
        
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            cell_type = CellType.EMPTY
        elif cell.ctype == xlrd.XL_CELL_TEXT:
            cell_type = CellType.STRING
        elif cell.ctype == xlrd.XL_CELL_NUMBER:
            cell_type = CellType.NUMBER
        elif cell.ctype == xlrd.XL_CELL_DATE:
            cell_type = CellType.DATE
            # 转换 Excel 日期为 Python datetime
            try:
                value = xlrd.xldate_as_datetime(value, workbook.datemode)
            except:
                pass
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            cell_type = CellType.BOOLEAN
            value = bool(value)
        elif cell.ctype == xlrd.XL_CELL_ERROR:
            cell_type = CellType.ERROR
        
        return CellData(
            value=value,
            cell_type=cell_type
        )
    
    @classmethod
    def format_file_size(cls, size_bytes: int) -> str:
        """格式化文件大小显示"""
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024 * 1024:
            return f"{size_bytes / 1024:.1f} KB"
        else:
            return f"{size_bytes / (1024 * 1024):.1f} MB"
