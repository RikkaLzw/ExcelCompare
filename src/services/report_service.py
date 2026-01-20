"""
报告生成服务

支持导出 Excel 和 HTML 格式的比较报告。
"""
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from src.models.excel_model import WorkbookData
from src.models.diff_model import CompareResult, DiffResult, DiffType


class ReportService:
    """报告生成服务"""
    
    # 差异类型对应的颜色
    DIFF_COLORS = {
        DiffType.MODIFIED: "FFFFC107",   # 黄色
        DiffType.ADDED: "FF4CAF50",      # 绿色
        DiffType.DELETED: "FFF44336",    # 红色
        DiffType.FORMAT_CHANGED: "FFFF9800",  # 橙色
    }
    
    @classmethod
    def export_excel(
        cls,
        result: CompareResult,
        workbook_a: WorkbookData,
        workbook_b: WorkbookData,
        output_path: str
    ):
        """
        导出 Excel 格式报告
        
        Args:
            result: 比较结果
            workbook_a: 工作簿 A
            workbook_b: 工作簿 B
            output_path: 输出路径
        """
        wb = openpyxl.Workbook()
        
        # 1. 摘要页
        ws_summary = wb.active
        ws_summary.title = "比较摘要"
        cls._write_summary_sheet(ws_summary, result, workbook_a, workbook_b)
        
        # 2. 差异详情页
        ws_details = wb.create_sheet("差异详情")
        cls._write_details_sheet(ws_details, result.diffs)
        
        # 3. 按工作表分页
        for sheet_name, diffs in result.diffs_by_sheet.items():
            if diffs:
                ws = wb.create_sheet(f"差异-{sheet_name}"[:31])  # 工作表名最长31字符
                cls._write_sheet_diffs(ws, sheet_name, diffs)
        
        wb.save(output_path)
    
    @classmethod
    def _write_summary_sheet(
        cls,
        ws,
        result: CompareResult,
        workbook_a: WorkbookData,
        workbook_b: WorkbookData
    ):
        """写入摘要页"""
        # 样式
        title_font = Font(size=16, bold=True)
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="E0E0E0")
        
        # 标题
        ws['A1'] = "Excel 文件比较报告"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # 比较信息
        ws['A3'] = "比较时间"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws['A4'] = "文件 A"
        ws['B4'] = workbook_a.file_name
        
        ws['A5'] = "文件 B"
        ws['B5'] = workbook_b.file_name
        
        # 统计信息
        ws['A7'] = "差异统计"
        ws['A7'].font = header_font
        ws.merge_cells('A7:D7')
        
        headers = ["类型", "数量", "占比"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        summary = result.summary
        total = summary.total if summary.total > 0 else 1  # 避免除零
        
        stats_data = [
            ("总计", summary.total, "100%"),
            ("修改", summary.modified, f"{summary.modified / total * 100:.1f}%"),
            ("新增", summary.added, f"{summary.added / total * 100:.1f}%"),
            ("删除", summary.deleted, f"{summary.deleted / total * 100:.1f}%"),
            ("格式变化", summary.format_changed, f"{summary.format_changed / total * 100:.1f}%"),
        ]
        
        for row, (type_name, count, ratio) in enumerate(stats_data, 9):
            ws.cell(row=row, column=1).value = type_name
            ws.cell(row=row, column=2).value = count
            ws.cell(row=row, column=3).value = ratio
        
        # 比较配置信息
        config = result.compare_config or {}
        if config:
            ws['A15'] = "比较配置"
            ws['A15'].font = header_font
            ws.merge_cells('A15:D15')
            
            config_row = 16
            if config.get('mode'):
                ws.cell(row=config_row, column=1).value = "比较模式"
                ws.cell(row=config_row, column=2).value = config.get('mode')
                config_row += 1
            
            if config.get('key_column') is not None:
                ws.cell(row=config_row, column=1).value = "主键列"
                ws.cell(row=config_row, column=2).value = f"第 {config.get('key_column') + 1} 列"
                config_row += 1
            
            if config.get('header_row') is not None:
                ws.cell(row=config_row, column=1).value = "标题行"
                ws.cell(row=config_row, column=2).value = f"第 {config.get('header_row') + 1} 行"
                config_row += 1
            
            if config.get('ignore_case'):
                ws.cell(row=config_row, column=1).value = "忽略大小写"
                ws.cell(row=config_row, column=2).value = "是"
                config_row += 1
            
            if config.get('ignore_whitespace'):
                ws.cell(row=config_row, column=1).value = "忽略空格"
                ws.cell(row=config_row, column=2).value = "是"
                config_row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
    
    @classmethod
    def _write_details_sheet(cls, ws, diffs: List[DiffResult]):
        """写入差异详情页"""
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="E0E0E0")
        
        headers = ["序号", "工作表", "位置", "类型", "原值", "新值"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        for row, diff in enumerate(diffs, 2):
            ws.cell(row=row, column=1).value = row - 1
            ws.cell(row=row, column=2).value = diff.sheet
            ws.cell(row=row, column=3).value = diff.position
            
            type_cell = ws.cell(row=row, column=4)
            type_cell.value = diff.type_display
            type_cell.fill = PatternFill("solid", fgColor=cls.DIFF_COLORS.get(diff.diff_type, "FFFFFF"))
            
            ws.cell(row=row, column=5).value = str(diff.old_value)[:1000] if diff.old_value else ""
            ws.cell(row=row, column=6).value = str(diff.new_value)[:1000] if diff.new_value else ""
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
    
    @classmethod
    def _write_sheet_diffs(cls, ws, sheet_name: str, diffs: List[DiffResult]):
        """写入单个工作表的差异"""
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="E0E0E0")
        
        ws['A1'] = f"工作表: {sheet_name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        headers = ["位置", "类型", "原值", "新值"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
        
        for row, diff in enumerate(diffs, 4):
            ws.cell(row=row, column=1).value = diff.position
            
            type_cell = ws.cell(row=row, column=2)
            type_cell.value = diff.type_display
            type_cell.fill = PatternFill("solid", fgColor=cls.DIFF_COLORS.get(diff.diff_type, "FFFFFF"))
            
            ws.cell(row=row, column=3).value = str(diff.old_value)[:500] if diff.old_value else ""
            ws.cell(row=row, column=4).value = str(diff.new_value)[:500] if diff.new_value else ""
    
    @classmethod
    def export_html(
        cls,
        result: CompareResult,
        workbook_a: WorkbookData,
        workbook_b: WorkbookData,
        output_path: str
    ):
        """
        导出 HTML 格式报告
        
        Args:
            result: 比较结果
            workbook_a: 工作簿 A
            workbook_b: 工作簿 B
            output_path: 输出路径
        """
        html_content = cls._generate_html(result, workbook_a, workbook_b)
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
    
    @classmethod
    def _generate_html(
        cls,
        result: CompareResult,
        workbook_a: WorkbookData,
        workbook_b: WorkbookData
    ) -> str:
        """生成 HTML 内容"""
        summary = result.summary
        
        # 差异行 HTML
        diff_rows = ""
        for i, diff in enumerate(result.diffs, 1):
            type_class = diff.diff_type.value
            old_val = str(diff.old_value)[:200] if diff.old_value else ""
            new_val = str(diff.new_value)[:200] if diff.new_value else ""
            diff_rows += f"""
            <tr class="{type_class}">
                <td>{i}</td>
                <td>{diff.sheet}</td>
                <td>{diff.position}</td>
                <td>{diff.type_display}</td>
                <td>{cls._escape_html(old_val)}</td>
                <td>{cls._escape_html(new_val)}</td>
            </tr>
            """
        
        html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel 比较报告</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        .card {{ background: white; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); margin-bottom: 20px; padding: 20px; }}
        h1 {{ color: #333; margin-bottom: 20px; }}
        h2 {{ color: #666; font-size: 18px; margin-bottom: 15px; }}
        .info-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px; }}
        .info-item {{ padding: 15px; background: #f8f8f8; border-radius: 6px; }}
        .info-label {{ font-size: 12px; color: #888; margin-bottom: 5px; }}
        .info-value {{ font-size: 16px; font-weight: 600; color: #333; }}
        .stats {{ display: flex; gap: 15px; flex-wrap: wrap; }}
        .stat-item {{ padding: 15px 20px; border-radius: 6px; text-align: center; min-width: 100px; }}
        .stat-value {{ font-size: 24px; font-weight: bold; }}
        .stat-label {{ font-size: 12px; color: #666; }}
        .total {{ background: #e3f2fd; color: #1976d2; }}
        .modified {{ background: #fff9c4; color: #f57c00; }}
        .added {{ background: #c8e6c9; color: #388e3c; }}
        .deleted {{ background: #ffcdd2; color: #d32f2f; }}
        .format {{ background: #ffe0b2; color: #e65100; }}
        table {{ width: 100%; border-collapse: collapse; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #eee; }}
        th {{ background: #f5f5f5; font-weight: 600; position: sticky; top: 0; }}
        tr:hover {{ background: #f8f8f8; }}
        tr.modified td:nth-child(4) {{ background: #fff9c4; }}
        tr.added td:nth-child(4) {{ background: #c8e6c9; }}
        tr.deleted td:nth-child(4) {{ background: #ffcdd2; }}
        tr.format td:nth-child(4) {{ background: #ffe0b2; }}
        .filter-bar {{ margin-bottom: 15px; display: flex; gap: 10px; align-items: center; }}
        .filter-bar input {{ padding: 8px 12px; border: 1px solid #ddd; border-radius: 4px; width: 200px; }}
        .filter-bar select {{ padding: 8px 12px; border: 1px solid #ddd; border-radius: 4px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="card">
            <h1>📊 Excel 文件比较报告</h1>
            <div class="info-grid">
                <div class="info-item">
                    <div class="info-label">比较时间</div>
                    <div class="info-value">{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">文件 A</div>
                    <div class="info-value">{workbook_a.file_name}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">文件 B</div>
                    <div class="info-value">{workbook_b.file_name}</div>
                </div>
                {cls._generate_config_html(result.compare_config)}
            </div>
        </div>
        
        <div class="card">
            <h2>差异统计</h2>
            <div class="stats">
                <div class="stat-item total">
                    <div class="stat-value">{summary.total}</div>
                    <div class="stat-label">总计</div>
                </div>
                <div class="stat-item modified">
                    <div class="stat-value">{summary.modified}</div>
                    <div class="stat-label">修改</div>
                </div>
                <div class="stat-item added">
                    <div class="stat-value">{summary.added}</div>
                    <div class="stat-label">新增</div>
                </div>
                <div class="stat-item deleted">
                    <div class="stat-value">{summary.deleted}</div>
                    <div class="stat-label">删除</div>
                </div>
                <div class="stat-item format">
                    <div class="stat-value">{summary.format_changed}</div>
                    <div class="stat-label">格式变化</div>
                </div>
            </div>
        </div>
        
        <div class="card">
            <h2>差异详情</h2>
            <div class="filter-bar">
                <input type="text" id="searchInput" placeholder="搜索..." onkeyup="filterTable()">
                <select id="typeFilter" onchange="filterTable()">
                    <option value="">全部类型</option>
                    <option value="modified">修改</option>
                    <option value="added">新增</option>
                    <option value="deleted">删除</option>
                    <option value="format">格式变化</option>
                </select>
            </div>
            <table id="diffTable">
                <thead>
                    <tr>
                        <th>序号</th>
                        <th>工作表</th>
                        <th>位置</th>
                        <th>类型</th>
                        <th>原值</th>
                        <th>新值</th>
                    </tr>
                </thead>
                <tbody>
                    {diff_rows}
                </tbody>
            </table>
        </div>
    </div>
    
    <script>
        function filterTable() {{
            const searchText = document.getElementById('searchInput').value.toLowerCase();
            const typeFilter = document.getElementById('typeFilter').value;
            const rows = document.querySelectorAll('#diffTable tbody tr');
            
            rows.forEach(row => {{
                const text = row.textContent.toLowerCase();
                const matchesSearch = text.includes(searchText);
                const matchesType = !typeFilter || row.classList.contains(typeFilter);
                row.style.display = matchesSearch && matchesType ? '' : 'none';
            }});
        }}
    </script>
</body>
</html>"""
        return html
    
    @staticmethod
    def _escape_html(text: str) -> str:
        """转义 HTML 特殊字符"""
        return (text
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;')
            .replace("'", '&#39;'))
    
    @classmethod
    def _generate_config_html(cls, config: dict) -> str:
        """生成比较配置的 HTML"""
        if not config:
            return ""
        
        items = []
        if config.get('mode'):
            items.append(f'''
                <div class="info-item">
                    <div class="info-label">比较模式</div>
                    <div class="info-value">{config.get('mode')}</div>
                </div>
            ''')
        
        if config.get('key_column') is not None:
            items.append(f'''
                <div class="info-item">
                    <div class="info-label">主键列</div>
                    <div class="info-value">第 {config.get('key_column') + 1} 列</div>
                </div>
            ''')
        
        if config.get('header_row') is not None:
            items.append(f'''
                <div class="info-item">
                    <div class="info-label">标题行</div>
                    <div class="info-value">第 {config.get('header_row') + 1} 行</div>
                </div>
            ''')
        
        # 忽略选项
        ignore_options = []
        if config.get('ignore_case'):
            ignore_options.append("大小写")
        if config.get('ignore_whitespace'):
            ignore_options.append("空格")
        if config.get('ignore_format'):
            ignore_options.append("格式")
        if config.get('ignore_empty_rows'):
            ignore_options.append("空行")
        
        if ignore_options:
            items.append(f'''
                <div class="info-item">
                    <div class="info-label">忽略选项</div>
                    <div class="info-value">{", ".join(ignore_options)}</div>
                </div>
            ''')
        
        # 选区信息
        if config.get('selection_a') and config.get('selection_b'):
            items.append(f'''
                <div class="info-item">
                    <div class="info-label">比较选区</div>
                    <div class="info-value">A: {config.get('selection_a')} ↔ B: {config.get('selection_b')}</div>
                </div>
            ''')
        
        return "".join(items)

